"""Flask web app — upload Excel, tìm URL khách sạn với realtime progress."""

import json
import os
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, jsonify, Response, send_file
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

from hotel_matcher import pick_url_for_hotel

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Lưu trữ progress của mỗi job: {job_id: {"status": "running/done/error", "rows": [...], "output": "..."}}
jobs = {}


def process_excel_job(job_id: str, input_path: str):
    """Xử lý Excel trong background thread, cập nhật progress vào jobs dict."""
    try:
        workbook = load_workbook(input_path)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        name_idx = headers.index("child_hotel_name") + 1
        addr_idx = headers.index("child_hotel_address") + 1

        url_col = len(headers) + 1
        engine_col = len(headers) + 2
        score_col = len(headers) + 3
        img_col = len(headers) + 4
        status_col = len(headers) + 5

        sheet.cell(row=1, column=url_col, value="official_website_url")
        sheet.cell(row=1, column=engine_col, value="search_engine_used")
        sheet.cell(row=1, column=score_col, value="match_score_address")
        sheet.cell(row=1, column=img_col, value="image_count")
        sheet.cell(row=1, column=status_col, value="status")

        total_rows = sheet.max_row - 1
        jobs[job_id]["total"] = total_rows

        with Stealth().use_sync(sync_playwright()) as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()

            for row in range(2, sheet.max_row + 1):
                hotel_name = (sheet.cell(row=row, column=name_idx).value or "").strip()
                hotel_address = (sheet.cell(row=row, column=addr_idx).value or "").strip()

                row_data = {
                    "row": row,
                    "hotel_name": hotel_name,
                    "hotel_address": hotel_address,
                    "status": "searching",
                    "url": "",
                    "score": 0,
                    "img_count": 0,
                }

                if not hotel_name or not hotel_address:
                    row_data["status"] = "missing-input"
                    sheet.cell(row=row, column=url_col, value="")
                    sheet.cell(row=row, column=engine_col, value="")
                    sheet.cell(row=row, column=score_col, value=0)
                    sheet.cell(row=row, column=img_col, value=0)
                    sheet.cell(row=row, column=status_col, value="missing-input")
                else:
                    url, engine, score, img_count, status = pick_url_for_hotel(page, hotel_name, hotel_address)
                    row_data["url"] = url
                    row_data["score"] = score
                    row_data["img_count"] = img_count
                    row_data["status"] = status

                    sheet.cell(row=row, column=url_col, value=url)
                    sheet.cell(row=row, column=engine_col, value=engine)
                    sheet.cell(row=row, column=score_col, value=score)
                    sheet.cell(row=row, column=img_col, value=img_count)
                    sheet.cell(row=row, column=status_col, value=status)

                jobs[job_id]["rows"].append(row_data)
                jobs[job_id]["current"] = row - 1

            context.close()
            browser.close()

        # Save output
        input_obj = Path(input_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(input_obj.parent / f"{input_obj.stem}_output_{ts}.xlsx")
        workbook.save(output_path)

        jobs[job_id]["status"] = "done"
        jobs[job_id]["output"] = output_path

    except Exception as e:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = str(e)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files supported"}), 400

    job_id = str(uuid.uuid4())[:8]
    save_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{job_id}_{file.filename}")
    file.save(save_path)

    jobs[job_id] = {
        "status": "running",
        "total": 0,
        "current": 0,
        "rows": [],
        "output": None,
        "error": None,
        "filename": file.filename,
    }

    thread = threading.Thread(target=process_excel_job, args=(job_id, save_path), daemon=True)
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/progress/<job_id>")
def progress(job_id):
    """SSE endpoint — stream progress realtime."""
    def generate():
        last_count = 0
        while True:
            if job_id not in jobs:
                yield f"data: {json.dumps({'error': 'Job not found'})}\n\n"
                break

            job = jobs[job_id]
            current_count = len(job["rows"])

            # Gửi các row mới
            if current_count > last_count:
                new_rows = job["rows"][last_count:current_count]
                for row in new_rows:
                    yield f"data: {json.dumps(row)}\n\n"
                last_count = current_count

            # Gửi status update
            yield f"data: {json.dumps({'type': 'status', 'status': job['status'], 'current': job.get('current', 0), 'total': job.get('total', 0)})}\n\n"

            if job["status"] in ("done", "error"):
                if job["status"] == "done":
                    yield f"data: {json.dumps({'type': 'complete', 'output': job['output']})}\n\n"
                else:
                    yield f"data: {json.dumps({'type': 'error', 'error': job.get('error', 'Unknown error')})}\n\n"
                break

            time.sleep(0.5)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/download/<job_id>")
def download(job_id):
    if job_id not in jobs or jobs[job_id]["status"] != "done":
        return jsonify({"error": "File not ready"}), 404

    output_path = jobs[job_id]["output"]
    return send_file(output_path, as_attachment=True,
                     download_name=f"output_{jobs[job_id]['filename']}")


if __name__ == "__main__":
    app.run(debug=True, port=5000, threaded=True)
