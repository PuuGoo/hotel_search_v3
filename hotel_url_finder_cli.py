"""CLI tool — tìm official website URL cho danh sách khách sạn từ Excel."""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

from hotel_matcher import pick_url_for_hotel


def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def emit_progress(data: dict, json_mode: bool):
    """Output progress — JSON mode: print JSON line, else: human readable."""
    if json_mode:
        print(json.dumps(data, ensure_ascii=False), flush=True)
    else:
        if data.get("type") == "row":
            status = data.get("status", "")
            name = data.get("hotel_name", "")
            url = data.get("url", "")
            log(f"  [{status}] {name} -> {url}")
        elif data.get("type") == "done":
            log(f"Done! Output: {data.get('output', '')}")


def process_excel(input_path: str, output_path: str | None = None, json_mode: bool = False) -> str:
    workbook = load_workbook(input_path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    name_idx = headers.index("child_hotel_name") + 1
    addr_idx = headers.index("child_hotel_address") + 1

    url_col = len(headers) + 1
    engine_col = len(headers) + 2
    score_col = len(headers) + 3
    img_col = len(headers) + 4
    status_col = len(headers) + 5

    sheet.cell(row=1, column=url_col, value="official_website_url")
    sheet.cell(row=1, column=engine_col, value="search_engine_used")
    sheet.cell(row=1, column=score_col, value="match_score_address")
    sheet.cell(row=1, column=img_col, value="image_count")
    sheet.cell(row=1, column=status_col, value="status")

    total_rows = sheet.max_row - 1

    emit_progress({"type": "start", "total": total_rows}, json_mode)

    with Stealth().use_sync(sync_playwright()) as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        for row in range(2, sheet.max_row + 1):
            hotel_name = (sheet.cell(row=row, column=name_idx).value or "").strip()
            hotel_address = (sheet.cell(row=row, column=addr_idx).value or "").strip()

            if not hotel_name or not hotel_address:
                sheet.cell(row=row, column=url_col, value="")
                sheet.cell(row=row, column=engine_col, value="")
                sheet.cell(row=row, column=score_col, value=0)
                sheet.cell(row=row, column=img_col, value=0)
                sheet.cell(row=row, column=status_col, value="missing-input")

                emit_progress({
                    "type": "row",
                    "row": row,
                    "total": total_rows,
                    "hotel_name": hotel_name,
                    "hotel_address": hotel_address,
                    "status": "missing-input",
                    "url": "",
                    "score": 0,
                    "img_count": 0,
                }, json_mode)
                continue

            url, engine, score, img_count, status = pick_url_for_hotel(page, hotel_name, hotel_address)
            sheet.cell(row=row, column=url_col, value=url)
            sheet.cell(row=row, column=engine_col, value=engine)
            sheet.cell(row=row, column=score_col, value=score)
            sheet.cell(row=row, column=img_col, value=img_count)
            sheet.cell(row=row, column=status_col, value=status)

            emit_progress({
                "type": "row",
                "row": row,
                "total": total_rows,
                "hotel_name": hotel_name,
                "hotel_address": hotel_address,
                "status": status,
                "url": url,
                "score": score,
                "img_count": img_count,
            }, json_mode)

        context.close()
        browser.close()

    # Save output
    if not output_path:
        input_obj = Path(input_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(input_obj.parent / f"{input_obj.stem}_output_{ts}.xlsx")

    workbook.save(output_path)

    emit_progress({"type": "done", "output": output_path, "total": total_rows}, json_mode)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Tìm official website URL cho danh sách khách sạn")
    parser.add_argument("--input", required=True, help="Đường dẫn file Excel đầu vào")
    parser.add_argument("--output", default=None, help="Đường dẫn file Excel đầu ra (tự动生成 nếu bỏ trống)")
    parser.add_argument("--json", action="store_true", help="Output progress dạng JSON (dùng cho web)")
    args = parser.parse_args()

    output = process_excel(args.input, args.output, args.json)
    if not args.json:
        print(output)


if __name__ == "__main__":
    main()
